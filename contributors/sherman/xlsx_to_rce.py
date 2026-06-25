#!/usr/bin/env python3
"""
xlsx_to_rce.py
==============
Convert Amanda's Gemini-extracted ALEKS XLSX into a valid RCEReport JSON
ready to drop straight into spot_check.html.

What gets extracted automatically
----------------------------------
- All subgroup findings + Prior Performance  (Subgroup Table sheet)
- Overall effect size, CI, n                 (first Grade Level row)
- Grade levels                               (Grade Level rows)
- Usage compliance rate                      (% Meeting Fidelity sheet)
- Average minutes on system                  (Usage Output sheet)
- Fidelity threshold                         (narrative text in % Meeting sheet)
- District name                              (School subgroup value, if present)
- Product name                               (inferred from workbook sheet names)

Provide the rest via CLI args (see --help).

Usage
-----
.venv/bin/python3 contributors/sherman/xlsx_to_rce.py \\
    "/path/to/ALEKS - Variable Mapping....xlsx" \\
    --period "08/25/2025 - 02/20/2026" \\
    --output data/annotations/aleks_from_gemini.json
"""

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(REPO_ROOT / "deliverables"))
from rce_schema import RCEReport  # noqa: E402

# ── Sheet fragments (case-insensitive partial match) ───────────────────────
SHEET_SUBGROUP = "subgroup table"
SHEET_FIDELITY = "% meeting"
SHEET_USAGE    = "usage output"
SHEET_PARAMS   = "basic study"

# ── Subgroup category → (schema_cat, custom_label) ────────────────────────
CAT_MAP = {
    "grade level":       ("grade_level",   None),
    "gender":            ("gender",        None),
    "ethnicity":         ("race_ethnicity", None),
    "school":            ("school",        None),
    "sped status":       ("custom",        "SpEd status"),
    "prior performance": ("custom",        "prior performance"),
    "ell":               ("ell",           None),
    "frl status":        ("frl_status",    None),
}


# ── Helpers ────────────────────────────────────────────────────────────────

def find_sheet(wb, fragment):
    frag = fragment.lower()
    for name in wb.sheetnames:
        if frag in name.lower():
            return wb[name]
    return None


def data_rows(ws):
    """Non-empty rows, header row skipped."""
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
    return rows[1:]


def text_rows(ws):
    """Yield non-empty string values from column A."""
    for row in ws.iter_rows(values_only=True):
        v = row[0] if row else None
        if v and isinstance(v, str):
            yield v.strip()


def map_significance(s):
    s = str(s or "").lower()
    if "not plotted" in s:
        return "not_reported"
    if "not statistically" in s or "undetermined" in s:
        return "not_significant"
    if "significant" in s:
        return "significant"
    return "not_reported"


def map_effectiveness(outcome_str):
    s = str(outcome_str or "").lower()
    if "better" in s or "positive" in s:
        return "positive"
    if "other students" in s or "negative" in s:
        return "negative"
    return "neutral"


def parse_ci(ci_str):
    nums = re.findall(r"-?\d+\.?\d*", str(ci_str or ""))
    if len(nums) >= 2:
        return float(nums[0]), float(nums[1])
    return None, None


def safe_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_grade(label):
    m = re.search(r"\d+", str(label or ""))
    return m.group() if m else str(label)


def derive_academic_year(period):
    years = re.findall(r"\d{4}", period)
    return f"{years[0]}-{years[1]}" if len(years) >= 2 else ""


def derive_period_label(period):
    m = re.search(r"(\d{1,2})/", period)
    start_month = int(m.group(1)) if m else 8
    return "BOY-MOY" if start_month >= 7 else "MOY-EOY"


def infer_product(wb):
    for name in wb.sheetnames:
        m = re.match(r"^([A-Z][A-Z0-9 ]+?)\s*[-–]", name)
        if m:
            candidate = m.group(1).strip()
            if candidate not in ("Verified", "Analysis", "Key", "Subgroup"):
                return candidate
    return "Unknown Product"


# ── Core extraction ────────────────────────────────────────────────────────

def extract(wb, args):
    subgroup_findings = []
    grade_levels_seen = set()
    overall_row = None  # (es, ci_lo, ci_hi, n, sig, eff)

    # ── Subgroup findings ─────────────────────────────────────────────────
    sub_ws = find_sheet(wb, SHEET_SUBGROUP)
    if not sub_ws:
        sys.exit("ERROR: Could not find 'Subgroup Table' sheet in the workbook.")

    current_cat_raw = None
    for row in data_rows(sub_ws):
        cat_cell = row[0]
        subgroup_val = row[1] if len(row) > 1 else None
        es_raw  = row[2] if len(row) > 2 else None
        ci_raw  = row[3] if len(row) > 3 else None
        n_raw   = row[4] if len(row) > 4 else None
        sig_raw = row[5] if len(row) > 5 else None
        out_raw = row[6] if len(row) > 6 else None

        if cat_cell is not None:
            current_cat_raw = str(cat_cell).strip()

        if not current_cat_raw or not subgroup_val:
            continue
        # Skip source/narrative rows
        if re.match(r"(source|http|note|\d+\.)", current_cat_raw, re.IGNORECASE):
            continue

        cat_key = current_cat_raw.lower()
        schema_cat, custom_label = CAT_MAP.get(cat_key, ("custom", current_cat_raw))

        not_plotted = str(es_raw or "").strip().lower() == "not plotted"
        es_val  = safe_float(es_raw) if not not_plotted else None
        ci_lo, ci_hi = parse_ci(ci_raw) if not not_plotted else (None, None)
        n_val   = int(float(n_raw)) if (not not_plotted and n_raw is not None) else None
        sig     = map_significance(sig_raw) if not not_plotted else "not_reported"
        eff     = map_effectiveness(out_raw)

        if schema_cat == "grade_level":
            grade_levels_seen.add(parse_grade(subgroup_val))
            if overall_row is None:
                overall_row = (es_val, ci_lo, ci_hi, n_val, sig, eff)
            continue  # Grade-level becomes the overall; not a subgroup row

        note = (
            f"Not plotted — {str(out_raw or '').strip() or 'sample size below statistical threshold (<30)'}"
            if not_plotted else None
        )
        subgroup_findings.append({
            "effectiveness": eff,
            "essa_tier": args.essa_tier,
            "subgroup_category": schema_cat,
            "subgroup_category_custom": custom_label,
            "subgroup_value": str(subgroup_val).strip(),
            "grade_levels": sorted(grade_levels_seen) or args.grade_levels,
            "usage_metric": "total_time",
            "outcome_measure": args.outcome,
            "effect_size_raw": str(es_raw).strip() if es_val is not None else None,
            "effect_size_value": es_val,
            "effect_size_type": "pearson_r",
            "statistical_significance": sig,
            "sample_size": n_val,
            "confidence_interval_lower": ci_lo,
            "confidence_interval_upper": ci_hi,
            "notes": note,
        })

    grade_levels = sorted(grade_levels_seen) if grade_levels_seen else args.grade_levels

    # Back-fill grade_levels for any findings added before the set was populated
    for sf in subgroup_findings:
        if not sf["grade_levels"]:
            sf["grade_levels"] = grade_levels

    # ── Overall effect size ───────────────────────────────────────────────
    if overall_row:
        ov_es, ov_ci_lo, ov_ci_hi, ov_n, ov_sig, ov_eff = overall_row
    else:
        ov_es = ov_ci_lo = ov_ci_hi = ov_n = None
        ov_sig, ov_eff = "not_reported", "not_reported"

    # ── Usage compliance ──────────────────────────────────────────────────
    usage_compliance_rate = None
    usage_baseline_n = None
    fidelity_threshold = args.threshold

    fid_ws = find_sheet(wb, SHEET_FIDELITY)
    if fid_ws:
        for row in data_rows(fid_ws):
            if str(row[0] or "").lower().startswith("grade level"):
                usage_baseline_n = int(float(row[3])) if row[3] is not None else None
                usage_compliance_rate = float(row[4]) if row[4] is not None else None
                break
        # Parse threshold from narrative (e.g. "the 665-minute usage benchmark")
        if not fidelity_threshold:
            for line in text_rows(fid_ws):
                m = re.search(r"(\d{3,4})[- ]minute", line, re.IGNORECASE)
                if m:
                    fidelity_threshold = f"{m.group(1)} Minutes on system"
                    break

    # ── Average minutes ───────────────────────────────────────────────────
    avg_minutes = None
    usage_ws = find_sheet(wb, SHEET_USAGE)
    if usage_ws:
        for row in data_rows(usage_ws):
            if str(row[0] or "").lower().startswith("grade level"):
                avg_minutes = int(float(row[2])) if row[2] is not None else None
                break

    # ── Fidelity goal from Basic Study Parameters ─────────────────────────
    if not fidelity_threshold:
        params_ws = find_sheet(wb, SHEET_PARAMS)
        if params_ws:
            for line in text_rows(params_ws):
                m = re.search(r"fidelity goal[:\s]+(.+)", line, re.IGNORECASE)
                if m:
                    fidelity_threshold = m.group(1).strip()
                    break

    # ── Usage notes ───────────────────────────────────────────────────────
    usage_notes = None
    if avg_minutes is not None:
        usage_notes = f"{avg_minutes} minutes average on system"
        if usage_baseline_n and ov_n:
            used_pct = round(usage_baseline_n / ov_n * 100)
            usage_notes += f"; {used_pct}% of students used the product"

    # ── Derived header fields ─────────────────────────────────────────────
    product = args.product or infer_product(wb)

    district = args.district
    if not district:
        for sf in subgroup_findings:
            if sf["subgroup_category"] == "school":
                district = sf["subgroup_value"]
                break
        if not district:
            district = "Unknown District"

    period = args.period
    academic_year = args.academic_year or derive_academic_year(period)
    period_label = args.period_label or derive_period_label(period)

    es_raw_str = None
    if ov_es is not None:
        es_raw_str = f"+{ov_es}" if ov_es > 0 else str(ov_es)

    return {
        "source_filename": args.xlsx.name,
        "extraction_model": "gemini (human-verified XLSX)",
        "extraction_timestamp": datetime.now(timezone.utc).isoformat(),
        "extraction_schema_version": "0.1",
        "report_title": f"{product} Usage & Outcomes Analysis — {period_label} {academic_year}, {district}",
        "product_name": product,
        "district_name": district,
        "academic_year": academic_year,
        "report_period": period,
        "report_period_label": period_label,
        "essa_tier": args.essa_tier,
        "essa_tier_label": f"Level {['I','II','III','IV'][args.essa_tier - 1]}",
        "analysis_method": "partial_correlation",
        "number_of_rces": 1,
        "covariates": ["school", "grade_level"],
        "grade_levels": grade_levels,
        "sample_size_max": ov_n,
        "outcome_measure": args.outcome,
        "overall_effectiveness": ov_eff,
        "overall_statistical_significance": ov_sig,
        "effect_size_raw": es_raw_str,
        "effect_size_value": ov_es,
        "effect_size_type": "pearson_r",
        "confidence_level": 0.95,
        "usage": {
            "usage_metric": "total_time",
            "usage_compliance_threshold": fidelity_threshold,
            "usage_compliance_rate": usage_compliance_rate,
            "usage_period_weeks": args.weeks,
            "usage_notes": usage_notes,
        },
        "subgroup_findings": subgroup_findings,
    }


# ── CLI ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Convert Gemini-extracted ALEKS XLSX to RCEReport JSON"
    )
    parser.add_argument("xlsx", type=Path, help="Path to the XLSX file")
    parser.add_argument("--product",      default=None,
                        help="Product name (default: inferred from sheet names)")
    parser.add_argument("--district",     default=None,
                        help="District name (default: inferred from School subgroup)")
    parser.add_argument("--period",       default="",
                        help='Report period e.g. "08/25/2025 - 02/20/2026"')
    parser.add_argument("--academic-year", dest="academic_year", default=None,
                        help="Academic year e.g. 2025-2026 (default: derived from period)")
    parser.add_argument("--period-label", dest="period_label", default=None,
                        help="BOY-MOY / MOY-EOY / Full Year (default: derived from period)")
    parser.add_argument("--outcome",      default="MOY FastBridge aMath Scores",
                        help='Primary outcome measure (default: "MOY FastBridge aMath Scores")')
    parser.add_argument("--essa-tier",    dest="essa_tier", type=int, default=3,
                        help="ESSA tier 1-4 (default: 3)")
    parser.add_argument("--grade-levels", dest="grade_levels", nargs="+", default=["6"],
                        help="Grade levels fallback if not found in sheet (default: 6)")
    parser.add_argument("--weeks",        type=int, default=None,
                        help="Study period in weeks (e.g. 25)")
    parser.add_argument("--threshold",    default=None,
                        help='Fidelity threshold e.g. "665 Minutes on system" (default: parsed from XLSX)')
    parser.add_argument("--no-validate",  dest="no_validate", action="store_true",
                        help="Skip Pydantic schema validation")
    parser.add_argument("--output", "-o", default=None,
                        help="Output JSON path (default: data/annotations/<xlsx_stem>.json)")
    args = parser.parse_args()

    if not args.xlsx.exists():
        sys.exit(f"ERROR: File not found: {args.xlsx}")

    print(f"Reading {args.xlsx.name} ...")
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    print(f"  Sheets: {', '.join(wb.sheetnames)}")

    result = extract(wb, args)

    if not args.no_validate:
        try:
            report = RCEReport(**result)
            result = json.loads(report.model_dump_json())
            print("Schema validation: OK")
        except Exception as e:
            print(f"Schema validation FAILED: {e}", file=sys.stderr)
            print("Writing raw output anyway.")

    out_path = (
        Path(args.output) if args.output
        else REPO_ROOT / "data" / "annotations" / f"{args.xlsx.stem}.json"
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(result, indent=2))

    subs = len(result.get("subgroup_findings") or [])
    print(f"Written to {out_path}")
    print(f"  {subs} subgroup findings · overall ES {result.get('effect_size_raw')} · n={result.get('sample_size_max')}")


if __name__ == "__main__":
    main()
